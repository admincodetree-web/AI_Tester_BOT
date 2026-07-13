import difflib
import json
import math
import os
import re
from collections import Counter
from html import unescape
from urllib.parse import urlparse, urljoin

import requests
from dotenv import load_dotenv
from flask import Flask, jsonify, render_template, request
from openpyxl import load_workbook

try:
    from spellchecker import SpellChecker
except ImportError:
    SpellChecker = None

load_dotenv()

app = Flask(__name__, template_folder='templates')

GEMINI_API_KEY = os.getenv('GEMINI_API_KEY')
GEMINI_MODEL = os.getenv('GEMINI_MODEL', 'gemini-1.5-flash')
GEMINI_USE_API = os.getenv('GEMINI_USE_API', 'false').strip().lower() in ('1', 'true', 'yes')

# OWASP ZAP configuration
ZAP_USE_API = os.getenv('ZAP_USE_API', 'false').strip().lower() in ('1', 'true', 'yes')
ZAP_API_KEY = os.getenv('ZAP_API_KEY', '')
ZAP_PROXY_URL = os.getenv('ZAP_PROXY_URL', 'http://127.0.0.1:8080')


def load_vulnerability_catalog():
    catalog_path = os.path.join(os.path.dirname(__file__), 'Web App Vulnerabilities.xlsx')
    if not os.path.exists(catalog_path):
        app.logger.warning('Vulnerability catalog file not found: %s', catalog_path)
        return []

    try:
        workbook = load_workbook(filename=catalog_path, read_only=True, data_only=True)
        sheet = workbook.active
        headers = [str(cell).strip() if cell else '' for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
        header_index = {name.lower(): idx for idx, name in enumerate(headers) if name}

        def get_value(row, key):
            idx = header_index.get(key)
            if idx is None or idx >= len(row):
                return ''
            value = row[idx]
            return str(value).strip() if value is not None else ''

        catalog = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            catalog.append({
                'id': get_value(row, '#') or get_value(row, 'id'),
                'severity': get_value(row, 'severity'),
                'vulnerabilityName': get_value(row, 'vulnerability name'),
                'details': get_value(row, 'vulnerability details'),
                'impact': get_value(row, 'impact'),
                'remediations': get_value(row, 'remediations'),
                'referenceLink': get_value(row, 'reference link'),
                'issueName': get_value(row, 'issue name'),
            })

        return catalog
    except Exception as exc:
        app.logger.warning('Unable to load vulnerability catalog: %s', exc)
        return []


VULNERABILITY_CATALOG = load_vulnerability_catalog()
CATALOG_NAMES = [item['vulnerabilityName'] for item in VULNERABILITY_CATALOG if item.get('vulnerabilityName')]


def is_valid_url(url: str) -> bool:
    try:
        parsed = urlparse(url.strip())
        return parsed.scheme in ('http', 'https') and bool(parsed.netloc)
    except Exception:
        return False


def fetch_url_content(url: str):
    """Fetch URL content with retries and graceful timeout handling.
    Returns: (page_content, response_headers, error_message)
    If connection fails, page_content and headers may be None but error_message explains why.
    """
    headers = {'User-Agent': 'AI-Security-Scanner/1.0'}
    last_error = None

    # Try with shorter timeout first, then fallback
    for timeout in [8, 12]:
        try:
            response = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
            response.raise_for_status()
            page_content = response.text[:16000]
            return page_content, dict(response.headers), None
        except requests.exceptions.Timeout:
            last_error = f'Connection timed out after {timeout}s. The target server may be down, blocking requests, or very slow.'
            app.logger.warning('Timeout fetching %s (timeout=%s)', url, timeout)
            continue
        except requests.exceptions.ConnectionError as exc:
            last_error = f'Connection failed: {str(exc)}. The server may be offline, blocking our scanner, or the URL may be incorrect.'
            app.logger.warning('Connection error fetching %s: %s', url, exc)
            break
        except requests.exceptions.HTTPError as exc:
            status = exc.response.status_code if exc.response else 'unknown'
            last_error = f'HTTP {status} error. The server responded with an error status code.'
            app.logger.warning('HTTP error fetching %s: %s', url, exc)
            # For 403/401, we still got headers — return them for analysis
            if exc.response and exc.response.status_code in (401, 403, 404, 500, 502, 503):
                return exc.response.text[:8000], dict(exc.response.headers), None
            break
        except Exception as exc:
            last_error = f'Unexpected error: {str(exc)}'
            app.logger.warning('Unexpected error fetching %s: %s', url, exc)
            break

    return None, None, last_error


def build_catalog_context() -> str:
    lines = []
    for item in VULNERABILITY_CATALOG:
        lines.append(f"- {item.get('vulnerabilityName', 'Unknown Vulnerability')}")
        if item.get('severity'):
            lines.append(f"  Severity: {item['severity']}")
        if item.get('details'):
            lines.append(f"  Details: {item['details'][:300]}")
        if item.get('referenceLink'):
            lines.append(f"  Reference: {item['referenceLink']}")
        if item.get('issueName'):
            lines.append(f"  Issue Name: {item['issueName']}")
        lines.append('')
    return '\n'.join(lines)


def build_gemini_prompt(url: str, headers: dict, page_content: str) -> str:
    catalog_context = build_catalog_context()
    return (
        'You are a security vulnerability analyzer. Here is a catalog of known vulnerabilities found in similar applications. '
        'Analyze the provided URL\'s content and headers, and identify which of these vulnerability patterns are present. '
        'ONLY report issues that have evidence in the content. Do NOT invent issues not in the catalog.\n\n'
        'Catalog:\n'
        f'{catalog_context}\n'
        f'URL: {url}\n\n'
        'Response Headers:\n'
        f'{json.dumps(headers, indent=2)}\n\n'
        'Page Content (truncated):\n'
        f'{page_content[:5000]}\n\n'
        'Identify potential security issues and return ONLY a JSON array in this exact format:\n'
        '[\n'
        '  {\n'
        '    "severity": "CRITICAL|HIGH|MEDIUM|LOW",\n'
        '    "vulnerabilityName": "Exact name from the catalog or closest match",\n'
        '    "referenceLink": "URL from the catalog or OWASP link",\n'
        '    "issueName": "Brief description of what was found",\n'
        '    "thinking": "Step-by-step reasoning with evidence from the content"\n'
        '  }\n'
        ']\n\n'
        'Return results in the exact JSON structure above and do not provide any other text outside the JSON array.\n'
    )


def extract_visible_text(html: str) -> str:
    if not html:
        return ''

    stripped = re.sub(r'<script[^>]*>.*?</script>', ' ', html, flags=re.DOTALL | re.IGNORECASE)
    stripped = re.sub(r'<style[^>]*>.*?</style>', ' ', stripped, flags=re.DOTALL | re.IGNORECASE)
    stripped = re.sub(r'<!--.*?-->', ' ', stripped, flags=re.DOTALL)
    stripped = re.sub(r'<[^>]+>', ' ', stripped)
    stripped = unescape(stripped)
    stripped = re.sub(r'\s+', ' ', stripped)
    return stripped.strip()


def find_aadhaar_patterns(text: str):
    if not text:
        return []

    patterns = [r'\b\d{4}\s?\d{4}\s?\d{4}\b', r'\b\d{12}\b']
    found = []
    for pattern in patterns:
        for match in re.findall(pattern, text):
            normalized = re.sub(r'\D', '', match)
            if normalized and normalized not in found:
                found.append(normalized)
    return found


def mask_aadhaar(value: str) -> str:
    cleaned = re.sub(r'\D', '', value)
    if len(cleaned) != 12:
        return value
    return f'XXXX-XXXX-{cleaned[-4:]}'


def detect_captcha(html: str):
    if not html:
        return []

    matches = []
    captcha_patterns = {
        'reCAPTCHA': r'recaptcha|g-recaptcha',
        'hCaptcha': r'hcaptcha',
        'Cloudflare Turnstile': r'cloudflare-turnstile',
        'FunCaptcha': r'funcaptcha',
        'SolveMedia': r'solvemedia',
    }
    for label, pattern in captcha_patterns.items():
        if re.search(pattern, html, re.IGNORECASE):
            matches.append(label)
    return matches


def check_login_form_security(url: str, html: str, headers: dict):
    findings = []
    if not html:
        return findings

    forms = re.findall(r'<form[^>]*>.*?</form>', html, flags=re.DOTALL | re.IGNORECASE)
    if not forms:
        return findings

    for form in forms:
        if not re.search(r'<input[^>]+type=["\']password["\']', form, re.IGNORECASE):
            continue

        method = re.search(r'method=["\']([^"\']+)["\']', form, re.IGNORECASE)
        action = re.search(r'action=["\']([^"\']+)["\']', form, re.IGNORECASE)
        autocomplete = re.search(r'autocomplete=["\']([^"\']+)["\']', form, re.IGNORECASE)

        method_value = method.group(1).strip().lower() if method else ''
        action_value = action.group(1).strip() if action else ''
        autocomplete_value = autocomplete.group(1).strip().lower() if autocomplete else ''

        insecure_reasons = []
        if method_value and method_value != 'post':
            insecure_reasons.append('form uses non-POST method')
        if action_value and action_value.startswith('http://'):
            insecure_reasons.append('form action uses http:// instead of https://')
        if autocomplete_value not in ('off', 'false'):
            insecure_reasons.append('autocomplete is not disabled for form credentials')

        if insecure_reasons:
            findings.append(build_static_result(
                'Login Form Security',
                'Insecure login form configuration detected: ' + '; '.join(insecure_reasons),
                'A password form was found and it contains one or more weak controls for secure credential handling. ' 
                'Ensure login forms use POST, HTTPS-based actions, and autocomplete off for credentials.',
                severity='HIGH',
            ))
        else:
            findings.append(build_static_result(
                'Login Form Validation',
                'Secure login form detected with secure form method and credential handling hints.',
                'A login form was found and appears to use standard best practices for secure credential submission.',
                severity='LOW',
            ))
        break

    return findings


def check_alignment_heuristics(html: str):
    findings = []
    if not html:
        return findings

    if re.search(r'<center\b|\balign=["\']', html, re.IGNORECASE):
        findings.append(build_static_result(
            'Alignment / Layout Practices',
            'Legacy alignment attributes or center tags were detected in the page markup.',
            'The page uses deprecated alignment techniques such as <center> or align="...". This can cause inconsistent layout across devices.',
            severity='LOW',
        ))

    layout_suspicions = re.findall(r'style=["\'][^"\']*(float:|position: absolute|position: fixed)[^"\']*["\']', html, re.IGNORECASE)
    if layout_suspicions:
        findings.append(build_static_result(
            'Alignment / Layout Check',
            'Inline positioning styles were found and may indicate fragile layout behavior.',
            'Inline CSS rules like float or absolute positioning can make page alignment inconsistent and difficult to maintain.',
            severity='LOW',
        ))

    return findings


def check_look_and_feel(html: str):
    findings = []
    if not html:
        return findings

    missing = []
    if not re.search(r'<title>.*?</title>', html, re.IGNORECASE):
        missing.append('title')
    if not re.search(r'<meta[^>]+name=["\']description["\']', html, re.IGNORECASE):
        missing.append('description meta tag')
    if not re.search(r'<meta[^>]+name=["\']viewport["\']', html, re.IGNORECASE):
        missing.append('viewport meta tag')
    if not re.search(r'<link[^>]+rel=["\'](?:icon|shortcut icon)["\']', html, re.IGNORECASE):
        missing.append('favicon link')
    if not re.search(r'<html[^>]+lang=["\'][a-zA-Z-]+["\']', html, re.IGNORECASE):
        missing.append('html lang attribute')

    if missing:
        findings.append(build_static_result(
            'Look and Feel Assessment',
            'Page quality checks are incomplete. Missing: ' + ', '.join(missing),
            'A professional page should include metadata such as title, description, viewport settings, favicon, and language attributes to improve accessibility and trust.',
            severity='LOW',
        ))

    return findings


def spell_check_findings(text: str):
    findings = []
    if not text:
        return findings

    words = re.findall(r"\b[a-zA-Z]{3,}\b", text)
    unique_words = list(dict.fromkeys(word.lower() for word in words))[:1000]
    misspelled = []

    if SpellChecker is not None:
        try:
            checker = SpellChecker()
            misspelled = list(checker.unknown(unique_words))[:10]
        except Exception:
            misspelled = []
    else:
        repeated = re.findall(r"\b([a-zA-Z]+)\s+\1\b", text, re.IGNORECASE)
        misspelled = [w.lower() for w in repeated][:10]

    if misspelled:
        suggestions = ', '.join(misspelled[:5])
        findings.append(build_static_result(
            'Spell Check',
            'Potential spelling or content issues detected: ' + suggestions,
            'The page contains text that may be misspelled or repeated. Content quality issues can reduce user trust and brand professionalism.',
            severity='LOW',
        ))

    return findings


def check_secure_layer_findings(url: str, headers: dict, html: str):
    findings = []
    lower_headers = {k.lower(): v for k, v in (headers or {}).items()}
    is_https = urlparse(url).scheme == 'https'

    if not is_https:
        findings.append(build_static_result(
            'Missing HTTPS Redirect',
            'The target URL uses http:// instead of https://.',
            'The connection is not encrypted by default, which can expose data in transit and make the application vulnerable to man-in-the-middle attacks.',
            severity='HIGH',
        ))
    elif 'strict-transport-security' not in lower_headers:
        findings.append(build_static_result(
            'Missing HSTS Header',
            'Strict-Transport-Security header is not present on the HTTPS response.',
            'HSTS helps prevent protocol downgrade attacks and ensures browsers only use HTTPS for the target domain.',
            severity='MEDIUM',
        ))

    if 'set-cookie' in lower_headers and 'secure' not in lower_headers.get('set-cookie', '').lower():
        findings.append(build_static_result(
            'Secure Cookie Flag Missing',
            'A cookie was set without the Secure flag.',
            'Cookies transmitted without the Secure attribute may be exposed when the application is served over HTTPS.',
            severity='MEDIUM',
        ))

    if is_https and re.search(r'src=["\']http://|href=["\']http://', html or '', re.IGNORECASE):
        findings.append(build_static_result(
            'Mixed Content Risk',
            'HTTP resources were referenced from an HTTPS page.',
            'Loading insecure assets on an HTTPS page can weaken the security of the entire page and expose it to interception.',
            severity='MEDIUM',
        ))

    return findings


def build_quality_findings(url: str, headers: dict, page_content: str):
    findings = []
    html = page_content or ''
    visible_text = extract_visible_text(html)

    findings.extend(check_look_and_feel(html))
    findings.extend(check_alignment_heuristics(html))
    findings.extend(check_login_form_security(url, html, headers))

    captcha_families = detect_captcha(html)
    if captcha_families:
        findings.append(build_static_result(
            'CAPTCHA Detection',
            'CAPTCHA protection detected: ' + ', '.join(captcha_families),
            'The page includes CAPTCHA or bot protection controls. This is useful for preventing automated abuse, but it is not a substitute for secure authentication or input validation.',
            severity='LOW',
        ))

    aadhaar_values = find_aadhaar_patterns(visible_text)
    if aadhaar_values:
        masked = ', '.join(mask_aadhaar(value) for value in aadhaar_values[:3])
        findings.append(build_static_result(
            'Aadhaar Data Exposure',
            f'Potential Aadhaar number(s) exposed in page text: {masked}',
            'The page appears to contain Indian Aadhaar numbers. Sensitive identifiers should be masked or removed from public pages to protect user privacy.',
            severity='HIGH',
        ))

    findings.extend(spell_check_findings(visible_text))
    findings.extend(check_secure_layer_findings(url, headers, html))
    return findings


def summarize_scan_results(results, url: str, headers: dict, page_content: str):
    counts = Counter((item.get('severity', 'LOW').upper() for item in results))
    total_items = sum(counts.values())
    html = page_content or ''
    look_and_feel_score = 0
    if re.search(r'<title>.*?</title>', html, re.IGNORECASE):
        look_and_feel_score += 20
    if re.search(r'<meta[^>]+name=["\']description["\']', html, re.IGNORECASE):
        look_and_feel_score += 20
    if re.search(r'<meta[^>]+name=["\']viewport["\']', html, re.IGNORECASE):
        look_and_feel_score += 20
    if re.search(r'<link[^>]+rel=["\'](?:icon|shortcut icon)["\']', html, re.IGNORECASE):
        look_and_feel_score += 20
    if re.search(r'<html[^>]+lang=["\'][a-zA-Z-]+["\']', html, re.IGNORECASE):
        look_and_feel_score += 20

    header_checks = ['x-frame-options', 'content-security-policy', 'strict-transport-security', 'x-content-type-options', 'referrer-policy']
    lower_headers = {k.lower(): v for k, v in (headers or {}).items()}
    header_score = sum(1 for h in header_checks if h in lower_headers)
    security_header_score = round((header_score / len(header_checks)) * 100) if header_checks else 0

    visible_text = extract_visible_text(html)
    misspelled = []
    if SpellChecker is not None:
        try:
            checker = SpellChecker()
            words = re.findall(r"\b[a-zA-Z]{3,}\b", visible_text)
            unique_words = list(dict.fromkeys(word.lower() for word in words))[:1000]
            misspelled = list(checker.unknown(unique_words))
        except Exception:
            misspelled = []

    aadhaar_values = find_aadhaar_patterns(visible_text)
    insecure_login = any(item.get('vulnerabilityName') == 'Login Form Security' for item in results)
    has_issue = total_items > 0

    quality_score = round(
        (look_and_feel_score * 0.3 + security_header_score * 0.3 + (0 if aadhaar_values else 100) * 0.2 + (0 if misspelled else 100) * 0.1 + (0 if insecure_login else 100) * 0.1) / 100
        * 100
    )
    quality_score = min(max(quality_score, 0), 100)

    return {
        'totals': {
            'critical': counts.get('CRITICAL', 0),
            'high': counts.get('HIGH', 0),
            'medium': counts.get('MEDIUM', 0),
            'low': counts.get('LOW', 0),
            'total': total_items,
        },
        'qualityScore': quality_score,
        'lookAndFeelScore': look_and_feel_score,
        'securityHeaderCoverage': security_header_score,
        'missingAadhaarMasks': len(aadhaar_values),
        'spellIssues': len(misspelled),
        'loginFormIssues': 1 if insecure_login else 0,
    }


def extract_token_counts(response, prompt: str, text: str):
    input_tokens = None
    output_tokens = None

    if hasattr(response, 'metadata'):
        metadata = getattr(response, 'metadata') or {}
    elif isinstance(response, dict):
        metadata = response.get('metadata', {}) or {}
    else:
        metadata = {}

    if isinstance(metadata, dict):
        input_tokens = metadata.get('input_tokens') or metadata.get('input_token_count') or metadata.get('inputTokenCount')
        output_tokens = metadata.get('output_tokens') or metadata.get('output_token_count') or metadata.get('outputTokenCount')
        token_usage = metadata.get('token_usage') or metadata.get('tokenUsage')
        if token_usage and isinstance(token_usage, dict):
            input_tokens = input_tokens or token_usage.get('input') or token_usage.get('input_tokens')
            output_tokens = output_tokens or token_usage.get('output') or token_usage.get('output_tokens')

    if input_tokens is None or output_tokens is None:
        input_tokens = max(1, len(prompt) // 4)
        output_tokens = max(1, len(text) // 4)
    try:
        return int(input_tokens), int(output_tokens)
    except Exception:
        return max(1, int(input_tokens or 1)), max(1, int(output_tokens or 1))


def format_token_annotation(input_tokens: int, output_tokens: int) -> str:
    return f'input={input_tokens} output={output_tokens}'


def extract_results_from_text(text: str):
    if not text:
        raise ValueError('Empty Gemini response')

    json_match = re.search(r'```json\s*(\[.*?\])\s*```', text, re.DOTALL | re.IGNORECASE)
    if json_match:
        payload = json_match.group(1)
    else:
        array_match = re.search(r'(\[\s*\{.*?\}\s*\])', text, re.DOTALL)
        payload = array_match.group(1) if array_match else text

    return json.loads(payload)


def normalize_results(results, token_annotation='input=0 output=0'):
    normalized = []
    for item in (results or []):
        severity = str(item.get('severity', 'LOW')).strip().upper()
        if severity not in ('CRITICAL', 'HIGH', 'MEDIUM', 'LOW'):
            severity = 'LOW'

        normalized.append({
            'severity': severity,
            'vulnerabilityName': str(item.get('vulnerabilityName', 'Unknown Vulnerability')).strip(),
            'referenceLink': str(item.get('referenceLink', 'https://owasp.org')).strip(),
            'issueName': str(item.get('issueName', 'No issue description provided')).strip(),
            'thinking': str(item.get('thinking', '')),
            'analysisTime': str(item.get('analysisTime', '')),
            'tokens': str(item.get('tokens', token_annotation)).strip(),
            'confidence': str(item.get('confidence', '')),
        })
    return normalized


def extract_text_from_gemini_response(response):
    if hasattr(response, 'text') and response.text is not None:
        return response.text
    if hasattr(response, 'output'):
        output = getattr(response, 'output')
        if isinstance(output, list) and output:
            first = output[0]
            if hasattr(first, 'content'):
                content = getattr(first, 'content')
                if isinstance(content, list) and content:
                    return getattr(content[0], 'text', str(content[0]))
            return getattr(first, 'text', str(first))
    if isinstance(response, dict):
        if 'text' in response:
            return response['text']
        if 'content' in response:
            return response['content']
        return json.dumps(response)
    return str(response)


def call_gemini_api(prompt: str):
    try:
        from google import genai
        from google.genai import types
    except ImportError as exc:
        raise RuntimeError(
            'Gemini integration requires the google GenAI package. Install with: pip install -r requirements.txt'
        ) from exc

    try:
        app.logger.info('Initializing Gemini client...')
        client = genai.Client(api_key=GEMINI_API_KEY)

        app.logger.info(f'Sending prompt to Gemini API ({GEMINI_MODEL})...')
        response = client.models.generate_content(
            model=GEMINI_MODEL,
            contents=prompt,
            config=types.GenerateContentConfig(
                temperature=0.0,
                top_p=1.0,
                top_k=1
            )
        )

        if not response:
            raise ValueError('Gemini API returned empty response')

        app.logger.info('Gemini API response received successfully')
        text = extract_text_from_gemini_response(response)

        if not text:
            raise ValueError('Failed to extract text from Gemini response')

        app.logger.info(f'Extracted {len(text)} characters from response')
        input_tokens, output_tokens = extract_token_counts(response, prompt, text)
        return text, input_tokens, output_tokens
    except Exception as exc:
        app.logger.error(f'Gemini API call failed: {type(exc).__name__}: {exc}', exc_info=True)
        raise


def find_catalog_entry_by_name(name: str):
    if not name:
        return None

    normalized_query = name.strip().lower()
    for item in VULNERABILITY_CATALOG:
        if item.get('vulnerabilityName', '').strip().lower() == normalized_query:
            return item

    if not CATALOG_NAMES:
        return None

    matches = difflib.get_close_matches(name, CATALOG_NAMES, n=1, cutoff=0.4)
    if matches:
        target = matches[0]
        return next((item for item in VULNERABILITY_CATALOG if item.get('vulnerabilityName') == target), None)

    return None


def build_static_result(name: str, issue_name: str, thinking: str, reference: str = None, severity: str = None):
    entry = find_catalog_entry_by_name(name)
    return {
        'severity': (severity or entry.get('severity', 'LOW')).strip().upper() if entry else (severity or 'LOW').strip().upper(),
        'vulnerabilityName': entry.get('vulnerabilityName') if entry else name,
        'referenceLink': reference or (entry.get('referenceLink') if entry else 'https://owasp.org'),
        'issueName': issue_name,
        'thinking': thinking,
        'analysisTime': '0.0s',
        'tokens': format_token_annotation(0, 0),
    }


def static_analysis_findings(url: str, headers: dict, page_content: str):
    findings = []

    if not headers:
        headers = {}
    if not page_content:
        page_content = ''

    lower_headers = {k.lower(): v for k, v in (headers or {}).items()}
    required_headers = [
        'x-frame-options',
        'content-security-policy',
        'strict-transport-security',
        'x-content-type-options',
        'referrer-policy',
    ]
    missing_headers = [h for h in required_headers if h not in lower_headers]
    if missing_headers:
        findings.append(build_static_result(
            'Missing Security Headers',
            f'Missing required security headers: {", ".join(missing_headers)}',
            'The response headers were inspected and one or more security headers are absent. '
            'This increases risk of clickjacking, MIME type confusion, and referrer leakage.',
        ))

    server_header = lower_headers.get('server')
    powered_by_header = lower_headers.get('x-powered-by')
    if server_header or powered_by_header:
        header_values = []
        if server_header:
            header_values.append(f'Server: {server_header}')
        if powered_by_header:
            header_values.append(f'X-Powered-By: {powered_by_header}')
        findings.append(build_static_result(
            'Information Disclosure',
            'Server version or platform disclosure found in response headers: ' + '; '.join(header_values),
            'The server response exposes implementation details that can help an attacker identify known vulnerabilities.',
        ))

    cors_origin = lower_headers.get('access-control-allow-origin', '')
    if cors_origin and cors_origin.strip() in ('*', 'null'):
        findings.append(build_static_result(
            'CORS Misconfiguration',
            f'Access-Control-Allow-Origin is set to {cors_origin!r}.',
            'A permissive CORS policy allows any origin to access resources, which may expose sensitive data in browser-based requests.',
        ))

    try:
        sensitive_pattern = re.search(r'\b(password|api_key|secret|token|encryptionkey)\b', page_content or '', re.IGNORECASE)
        if sensitive_pattern:
            findings.append(build_static_result(
                'Information Disclosure',
                f'Potential sensitive data exposure found by matching {sensitive_pattern.group(0)} in page content.',
                'The page content contains terms commonly associated with credentials or secrets, which may indicate accidental disclosure.',
            ))
    except Exception as e:
        app.logger.warning(f'Sensitive pattern search failed: {e}')

    try:
        parsed_url = urlparse(url)
        if parsed_url.scheme == 'http':
            findings.append(build_static_result(
                'Missing HTTPS Redirect',
                'The target URL uses http:// instead of https://.',
                'The connection is not encrypted by default, which can expose data in transit and make the application vulnerable to man-in-the-middle attacks.',
            ))
    except Exception as e:
        app.logger.warning(f'URL parsing failed: {e}')

    findings.extend(build_quality_findings(url, headers, page_content))
    return findings


def analyze_unreachable_url(url: str, error_msg: str):
    """When a URL cannot be fetched, still perform analysis on what we know."""
    findings = []

    # Check if it's HTTP
    try:
        parsed = urlparse(url)
        if parsed.scheme == 'http':
            findings.append(build_static_result(
                'Missing HTTPS Redirect',
                'The target URL uses http:// instead of https://.',
                'The connection is not encrypted by default. Additionally, the server could not be reached, which may indicate it is down or blocking requests.',
            ))
    except Exception:
        pass

    # Add a note about connectivity
    findings.append(build_static_result(
        'Information Disclosure',
        f'Target server could not be reached: {error_msg}',
        'The scanner was unable to establish a connection to the target. This could mean the server is offline, blocking automated requests, or the URL is incorrect. '
        'No content or headers could be analyzed. Consider testing with a different URL or checking if the target requires authentication.',
        severity='LOW',
    ))

    return findings


def run_zap_scan(target_url: str):
    """Connect to OWASP ZAP daemon and run a Spider + Passive Scan.
    Returns a list of normalized findings on success, or an empty list on failure.
    ZAP must be running at ZAP_PROXY_URL with API access enabled.
    Active scan is intentionally skipped to avoid long timeouts in serverless environments.
    """
    if not ZAP_USE_API or not ZAP_API_KEY:
        return []

    try:
        import time
        from zapv2 import ZAPv2
    except ImportError:
        app.logger.warning('python-owasp-zap-v2.4 not installed. Skipping ZAP scan.')
        return []

    try:
        app.logger.info(f'Connecting to ZAP at {ZAP_PROXY_URL}...')
        zap = ZAPv2(
            proxies={'http': ZAP_PROXY_URL, 'https': ZAP_PROXY_URL},
            apikey=ZAP_API_KEY,
        )

        # Verify ZAP is reachable
        version = zap.core.version
        app.logger.info(f'ZAP connection successful. Version: {version}')

        # Access the target URL so ZAP registers it in its site tree
        zap.core.access_url(target_url, followredirects=True)
        app.logger.info(f'ZAP accessed target URL: {target_url}')

        # Run Spider to crawl accessible pages (passive, no payload injection)
        spider_id = zap.spider.scan(url=target_url, apikey=ZAP_API_KEY)
        app.logger.info(f'ZAP spider started, ID: {spider_id}')

        # Poll spider with a max wait of 30 seconds (safe for Vercel Pro)
        max_spider_wait = 30
        elapsed = 0
        while int(zap.spider.status(spider_id)) < 100 and elapsed < max_spider_wait:
            time.sleep(2)
            elapsed += 2

        spider_status = zap.spider.status(spider_id)
        app.logger.info(f'ZAP spider finished at {spider_status}% after {elapsed}s.')

        # Wait briefly for passive scan queue to process
        time.sleep(3)

        # Retrieve alerts (passive scan findings)
        alerts = zap.core.alerts(baseurl=target_url)
        app.logger.info(f'ZAP returned {len(alerts)} alerts.')

        findings = []
        for alert in alerts:
            raw_risk = alert.get('risk', 'Low')
            # Map ZAP risk levels to our severity format
            severity_map = {'High': 'HIGH', 'Medium': 'MEDIUM', 'Low': 'LOW', 'Informational': 'LOW'}
            severity = severity_map.get(raw_risk, 'LOW')

            ref = alert.get('reference', '').strip()
            # ZAP references can be multi-line; take the first URL-like line
            ref_url = next(
                (line.strip() for line in ref.splitlines() if line.strip().startswith('http')),
                'https://owasp.org'
            )

            evidence = alert.get('evidence', '')
            solution = alert.get('solution', '')
            description = alert.get('description', '')
            thinking = f"Description: {description}\nEvidence: {evidence}\nSolution: {solution}".strip()

            findings.append({
                'severity': severity,
                'vulnerabilityName': alert.get('alert', 'Unknown ZAP Finding'),
                'referenceLink': ref_url,
                'issueName': alert.get('name', alert.get('alert', 'No description')),
                'thinking': thinking,
                'analysisTime': f'{elapsed}s (ZAP Spider)',
                'tokens': format_token_annotation(0, 0),
                'source': 'ZAP',
            })

        return findings

    except Exception as exc:
        app.logger.error(f'ZAP scan failed: {type(exc).__name__}: {exc}', exc_info=True)
        return []


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/scan', methods=['POST'])
def scan():
    try:
        data = request.get_json(silent=True) or {}
        url = data.get('url', '').strip()

        if not url or not is_valid_url(url):
            return jsonify({'error': 'Please provide a valid http or https URL.', 'results': []}), 400

        page_content, headers, fetch_error = fetch_url_content(url)

        # If fetch failed completely, still try to analyze what we can from the URL itself
        if fetch_error and (page_content is None or headers is None):
            app.logger.warning('Fetch failed for %s: %s', url, fetch_error)
            fallback_results = analyze_unreachable_url(url, fetch_error)
            return jsonify({
                'results': fallback_results,
                'url': url,
                'warning': f'Could not fetch target content: {fetch_error}. Analysis is limited to URL-level checks only.',
            })

        try:
            static_results = static_analysis_findings(url, headers or {}, page_content or '')
            app.logger.info(f'Static analysis found {len(static_results)} issues')
        except Exception as exc:
            app.logger.error(f'Static analysis error: {exc}', exc_info=True)
            static_results = []

        ai_results = []

        if GEMINI_USE_API and GEMINI_API_KEY and page_content:
            prompt = build_gemini_prompt(url, headers or {}, page_content)
            try:
                app.logger.info('Calling Gemini API...')
                raw_text, input_tokens, output_tokens = call_gemini_api(prompt)
                app.logger.info(f'Gemini response received: {len(raw_text)} chars')
                extracted = extract_results_from_text(raw_text)
                app.logger.info(f'Extracted {len(extracted)} AI results')
                ai_results = normalize_results(
                    extracted,
                    format_token_annotation(input_tokens, output_tokens),
                )
            except Exception as exc:
                app.logger.error(f'Gemini API error: {exc}', exc_info=True)
        elif GEMINI_USE_API and not GEMINI_API_KEY:
            app.logger.warning('GEMINI_USE_API is enabled but GEMINI_API_KEY is not set. Skipping AI request.')

        # OWASP ZAP active/passive scan integration
        zap_results = []
        if ZAP_USE_API and ZAP_API_KEY:
            app.logger.info('Starting OWASP ZAP scan...')
            zap_results = run_zap_scan(url)
            app.logger.info(f'ZAP scan returned {len(zap_results)} findings.')
        elif ZAP_USE_API and not ZAP_API_KEY:
            app.logger.warning('ZAP_USE_API is enabled but ZAP_API_KEY is not set. Skipping ZAP scan.')

        combined = []
        seen = set()
        # Merge: static checks + Gemini AI findings + ZAP findings
        for result in static_results + ai_results + zap_results:
            key = result.get('vulnerabilityName', '').strip().lower()
            if not key:
                continue
            if key not in seen:
                combined.append(result)
                seen.add(key)

        summary = summarize_scan_results(combined, url, headers or {}, page_content or '')

        if not combined:
            return jsonify({
                'results': [],
                'url': url,
                'summary': summary,
                'message': 'No vulnerabilities detected in the accessible content. Note: This tool can only analyze publicly visible content and headers. Active vulnerabilities (SQL injection, broken access control, IDOR, etc.) require authenticated penetration testing and cannot be detected through static analysis.',
            })

        return jsonify({'results': combined, 'url': url, 'summary': summary})

    except Exception as exc:
        app.logger.error(f'Unexpected error in /scan: {exc}', exc_info=True)
        return jsonify({'error': f'Internal server error: {str(exc)}', 'results': []}), 500


if __name__ == '__main__':
    port = int(os.getenv('PORT', '5000'))
    debug = os.getenv('FLASK_DEBUG', 'false').strip().lower() in ('1', 'true', 'yes')
    app.run(debug=debug, host='0.0.0.0', port=port)
